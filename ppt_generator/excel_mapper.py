"""Generic, UI-independent Excel range preview and PPT Slot mapping."""

from __future__ import annotations

import re
from dataclasses import asdict, dataclass
from datetime import date, datetime, time
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries


PREVIEW_MAX_ROWS = 300
PREVIEW_MAX_COLUMNS = 80


@dataclass(frozen=True)
class ExcelSheetPreview:
    name: str
    values: tuple[tuple[str, ...], ...]
    source_max_row: int
    source_max_column: int
    truncated: bool = False

    @property
    def row_count(self) -> int:
        return len(self.values)

    @property
    def column_count(self) -> int:
        return max((len(row) for row in self.values), default=0)


@dataclass(frozen=True)
class ExcelWorkbookPreview:
    path: Path
    sheets: tuple[ExcelSheetPreview, ...]

    def sheet(self, name: str) -> ExcelSheetPreview:
        for sheet in self.sheets:
            if sheet.name == name:
                return sheet
        raise ValueError(f"Excel 中不存在工作表：{name}")


@dataclass
class ExcelMappingRule:
    sheet: str
    source_range: str
    target_slot: str
    mode: str = "text"
    separator: str = "\n"
    enabled: bool = True

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    @classmethod
    def from_dict(cls, raw: Mapping[str, Any]) -> "ExcelMappingRule":
        rule = cls(
            sheet=str(raw.get("sheet") or ""),
            source_range=str(raw.get("source_range") or "").upper(),
            target_slot=str(raw.get("target_slot") or ""),
            mode=str(raw.get("mode") or "text"),
            separator=str(raw.get("separator") if raw.get("separator") is not None else "\n"),
            enabled=bool(raw.get("enabled", True)),
        )
        rule.validate()
        return rule

    def validate(self) -> None:
        if not self.sheet:
            raise ValueError("Excel 映射缺少工作表")
        if not self.source_range:
            raise ValueError("Excel 映射缺少来源范围")
        try:
            range_boundaries(self.source_range)
        except ValueError as exc:
            raise ValueError(f"Excel 范围无效：{self.source_range}") from exc
        if not self.target_slot:
            raise ValueError("Excel 映射缺少目标 Slot")
        if self.mode not in {"text", "table"}:
            raise ValueError(f"Excel 映射模式不支持：{self.mode}")


@dataclass(frozen=True)
class MappingTarget:
    key: str
    label: str
    kind: str = "text"


@dataclass(frozen=True)
class MappingApplicationResult:
    values: dict[str, Any]
    warnings: tuple[str, ...]


def excel_value_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _trim_preview(rows: list[list[str]]) -> tuple[tuple[str, ...], ...]:
    while rows and not any(cell for cell in rows[-1]):
        rows.pop()
    max_used_column = 0
    for row in rows:
        for index, cell in enumerate(row, start=1):
            if cell:
                max_used_column = max(max_used_column, index)
    if not max_used_column:
        return ()
    return tuple(tuple(row[:max_used_column]) for row in rows)


def load_excel_preview(
    path: str | Path,
    *,
    max_rows: int = PREVIEW_MAX_ROWS,
    max_columns: int = PREVIEW_MAX_COLUMNS,
) -> ExcelWorkbookPreview:
    source = Path(path).expanduser().resolve()
    if source.suffix.lower() not in {".xlsx", ".xlsm"} or not source.is_file():
        raise FileNotFoundError(f"Excel 文件不存在或格式不支持：{source}")
    workbook = load_workbook(source, read_only=True, data_only=True)
    previews: list[ExcelSheetPreview] = []
    try:
        for worksheet in workbook.worksheets:
            source_rows = int(worksheet.max_row or 0)
            source_columns = int(worksheet.max_column or 0)
            preview_rows = min(source_rows, max_rows)
            preview_columns = min(source_columns, max_columns)
            values = [
                [excel_value_text(value) for value in row]
                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=max(1, preview_rows),
                    min_col=1,
                    max_col=max(1, preview_columns),
                    values_only=True,
                )
            ] if source_rows and source_columns else []
            previews.append(
                ExcelSheetPreview(
                    name=worksheet.title,
                    values=_trim_preview(values),
                    source_max_row=source_rows,
                    source_max_column=source_columns,
                    truncated=source_rows > max_rows or source_columns > max_columns,
                )
            )
    finally:
        workbook.close()
    return ExcelWorkbookPreview(source, tuple(previews))


def detect_header_row(sheet: ExcelSheetPreview, *, scan_rows: int = 30) -> int:
    """Return a conservative one-based header-row suggestion."""
    best_row = 1
    best_score = -1.0
    for index, row in enumerate(sheet.values[:scan_rows], start=1):
        nonempty = [cell.strip() for cell in row if cell.strip()]
        if len(nonempty) < 2:
            continue
        text_cells = sum(not _looks_numeric(cell) for cell in nonempty)
        next_nonempty = 0
        if index < len(sheet.values):
            next_nonempty = sum(bool(cell.strip()) for cell in sheet.values[index])
        score = len(nonempty) * 2 + text_cells + min(next_nonempty, len(nonempty)) * 0.25
        if score > best_score:
            best_row = index
            best_score = score
    return best_row


def _looks_numeric(value: str) -> bool:
    try:
        float(value.replace(",", ""))
        return True
    except ValueError:
        return False


def range_shape(source_range: str) -> tuple[int, int]:
    min_column, min_row, max_column, max_row = range_boundaries(source_range)
    return max_row - min_row + 1, max_column - min_column + 1


def selection_range(
    min_row: int,
    min_column: int,
    max_row: int,
    max_column: int,
) -> str:
    start = f"{get_column_letter(min_column)}{min_row}"
    end = f"{get_column_letter(max_column)}{max_row}"
    return start if start == end else f"{start}:{end}"


def _matrix_from_worksheet(worksheet: Any, source_range: str) -> list[list[str]]:
    min_column, min_row, max_column, max_row = range_boundaries(source_range)
    if min_row < 1 or min_column < 1:
        raise ValueError(f"Excel 范围无效：{source_range}")
    return [
        [excel_value_text(value) for value in row]
        for row in worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_column,
            max_col=max_column,
            values_only=True,
        )
    ]


def matrix_to_text(matrix: list[list[str]], separator: str = "\n") -> str:
    lines: list[str] = []
    for row in matrix:
        cells = list(row)
        while cells and not cells[-1].strip():
            cells.pop()
        if not cells:
            continue
        lines.append("\t".join(cells))
    return separator.join(lines)


def preview_rule_value(sheet: ExcelSheetPreview, rule: ExcelMappingRule) -> Any:
    rule.validate()
    min_column, min_row, max_column, max_row = range_boundaries(rule.source_range)
    if max_row > sheet.row_count or max_column > sheet.column_count:
        return "预览范围外"
    matrix = [
        list(sheet.values[row - 1][min_column - 1 : max_column])
        for row in range(min_row, max_row + 1)
    ]
    return matrix if rule.mode == "table" else matrix_to_text(matrix, rule.separator)


def apply_excel_mappings(
    path: str | Path,
    rules: Iterable[ExcelMappingRule],
    *,
    slot_specs: Mapping[str, Mapping[str, Any]] | None = None,
) -> MappingApplicationResult:
    source = Path(path).expanduser().resolve()
    enabled_rules = [rule for rule in rules if rule.enabled]
    targets = [rule.target_slot for rule in enabled_rules]
    duplicate_targets = sorted({target for target in targets if targets.count(target) > 1})
    if duplicate_targets:
        raise ValueError(f"同一目标存在多条启用映射：{', '.join(duplicate_targets)}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    values: dict[str, Any] = {}
    warnings: list[str] = []
    try:
        for rule in enabled_rules:
            rule.validate()
            if slot_specs is not None and rule.target_slot not in slot_specs:
                raise ValueError(f"PPT 模板不存在目标 Slot：{rule.target_slot}")
            if rule.sheet not in workbook.sheetnames:
                raise ValueError(f"Excel 中不存在工作表：{rule.sheet}")
            worksheet = workbook[rule.sheet]
            matrix = _matrix_from_worksheet(worksheet, rule.source_range)
            if rule.mode == "table":
                if slot_specs is not None:
                    spec = slot_specs[rule.target_slot]
                    expected = (spec.get("rows"), spec.get("columns"))
                    actual = (len(matrix), len(matrix[0]) if matrix else 0)
                    if all(isinstance(item, int) for item in expected) and actual != expected:
                        raise ValueError(
                            f"{rule.target_slot} 需要 {expected[0]}×{expected[1]}，"
                            f"Excel 范围 {rule.source_range} 为 {actual[0]}×{actual[1]}"
                        )
                values[rule.target_slot] = matrix
            else:
                text = matrix_to_text(matrix, rule.separator)
                values[rule.target_slot] = text
                if not text.strip():
                    warnings.append(
                        f"{rule.sheet}!{rule.source_range} 为空，目标 {rule.target_slot} 将写入空文字"
                    )
    finally:
        workbook.close()
    return MappingApplicationResult(values, tuple(warnings))


def _normalize_label(value: str) -> str:
    return re.sub(r"[\s\W_]+", "", value, flags=re.UNICODE).lower()


def _label_score(source: str, target: MappingTarget) -> float:
    source_norm = _normalize_label(source)
    options = [_normalize_label(target.label), _normalize_label(target.key)]
    best = 0.0
    for option in options:
        if not source_norm or not option:
            continue
        if source_norm == option:
            return 1.0
        if min(len(source_norm), len(option)) >= 3 and (
            source_norm in option or option in source_norm
        ):
            best = max(best, 0.9)
        best = max(best, SequenceMatcher(None, source_norm, option).ratio())
    return best


def suggest_text_mappings(
    sheet: ExcelSheetPreview,
    targets: Iterable[MappingTarget],
    *,
    threshold: float = 0.74,
) -> list[ExcelMappingRule]:
    """Suggest adjacent key-value mappings; never applies them automatically."""
    available_targets = [target for target in targets if target.kind == "text"]
    suggestions: list[ExcelMappingRule] = []
    used_targets: set[str] = set()
    for row_index, row in enumerate(sheet.values[:100], start=1):
        for column_index, cell in enumerate(row[:20], start=1):
            if not cell.strip() or column_index >= len(row):
                continue
            right_value = row[column_index]
            if not right_value.strip():
                continue
            candidates = [
                (target, _label_score(cell, target))
                for target in available_targets
                if target.key not in used_targets
            ]
            if not candidates:
                continue
            target, score = max(candidates, key=lambda item: item[1])
            if score < threshold:
                continue
            suggestions.append(
                ExcelMappingRule(
                    sheet=sheet.name,
                    source_range=f"{get_column_letter(column_index + 1)}{row_index}",
                    target_slot=target.key,
                    mode="text",
                )
            )
            used_targets.add(target.key)
    return suggestions
