"""Project-module operations and structure planning, independent from PySide6."""

from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from .excel_mapper import excel_value_text
from .project import (
    ExcelModuleBinding,
    PageTemplateRef,
    PptProject,
    ProjectModule,
    ProjectSlide,
    new_project_id,
)

if TYPE_CHECKING:
    from .template_renderer import TemplateManifest


MODULE_TYPES = {"fixed", "repeat", "dataDriven"}
PAGE_ROLES = {"cover", "section", "content", "parameters", "gallery", "ending"}
NUMBER_EXCLUDED_ROLES = {"cover", "ending"}


@dataclass(frozen=True)
class SlideContext:
    module_id: str
    slide_id: str
    template_module_key: str
    page_template_key: str
    source_slide: int
    role: str
    physical_page_number: int
    physical_total_pages: int
    page_number: int | None
    total_pages: int
    module_number: int
    module_count: int
    module_page_number: int
    module_page_count: int
    module_title: str
    slide_title: str
    slide_subtitle: str
    values: dict[str, Any]
    system_slots: dict[str, int]
    remove_shape_ids: list[int]
    number_format: str


def _module_role(module_key: str, explicit_role: str = "") -> str:
    if explicit_role in PAGE_ROLES:
        return explicit_role
    lowered = module_key.lower()
    if lowered in {"cover", "front_cover"}:
        return "cover"
    if lowered in {"ending", "end", "back_cover"}:
        return "ending"
    return "content"


def _page_templates_from_manifest_module(
    module: dict[str, Any],
    *,
    slide_count: int,
) -> tuple[list[PageTemplateRef], list[str], str]:
    module_key = str(module["key"])
    module_name = str(module.get("name") or module_key)
    raw_templates = module.get("page_templates")
    templates: list[PageTemplateRef] = []
    if isinstance(raw_templates, list) and raw_templates:
        for index, raw in enumerate(raw_templates, start=1):
            if not isinstance(raw, dict):
                raise ValueError(f"模块 {module_name} 的页面模板必须是对象")
            source_slide = int(raw.get("source_slide") or 0)
            if not 1 <= source_slide <= slide_count:
                raise ValueError(f"模块 {module_name} 的页面模板页码无效：{source_slide}")
            key = str(raw.get("key") or f"{module_key}_page_{source_slide}_{index}")
            templates.append(
                PageTemplateRef(
                    key=key,
                    name=str(raw.get("name") or f"{module_name} · 模板第{source_slide}页"),
                    source_slide=source_slide,
                    role=_module_role(module_key, str(raw.get("role") or "")),
                    default_title=str(raw.get("default_title") or ""),
                    default_subtitle=str(raw.get("default_subtitle") or ""),
                    system_slots={
                        str(name): int(shape_id)
                        for name, shape_id in dict(raw.get("system_slots") or {}).items()
                    },
                    remove_shape_ids=[
                        int(shape_id) for shape_id in raw.get("remove_shape_ids", [])
                    ],
                    number_format=str(
                        raw.get("number_format") or "{page_number} / {total_pages}"
                    ),
                )
            )
    else:
        slides = module.get("slides") or []
        for index, source_slide in enumerate(slides, start=1):
            source_slide = int(source_slide)
            page_name = (
                module_name
                if len(slides) == 1
                else f"{module_name} · 第{index}页"
            )
            templates.append(
                PageTemplateRef(
                    key=f"{module_key}_page_{source_slide}",
                    name=page_name,
                    source_slide=source_slide,
                    role=_module_role(module_key),
                    default_title=page_name,
                )
            )

    keys = [item.key for item in templates]
    if len(keys) != len(set(keys)):
        raise ValueError(f"模块 {module_name} 的页面模板 key 重复")
    requested_sequence = module.get("default_sequence")
    default_sequence = (
        [str(item) for item in requested_sequence]
        if isinstance(requested_sequence, list) and requested_sequence
        else keys[:]
    )
    unknown = [key for key in default_sequence if key not in set(keys)]
    if unknown:
        raise ValueError(f"模块 {module_name} 的默认页面序列含未知模板：{', '.join(unknown)}")
    requested_default = str(module.get("default_add_template") or "")
    if requested_default and requested_default not in set(keys):
        raise ValueError(f"模块 {module_name} 的默认新增页面模板不存在")
    default_add_template = requested_default or (keys[0] if len(keys) == 1 else "")
    return templates, default_sequence, default_add_template


def _new_module_from_manifest(module: dict[str, Any], slide_count: int) -> ProjectModule:
    templates, default_sequence, default_add_template = _page_templates_from_manifest_module(
        module, slide_count=slide_count
    )
    templates_by_key = {item.key: item for item in templates}
    slides = [
        ProjectSlide(
            page_template_key=key,
            title=templates_by_key[key].default_title or templates_by_key[key].name,
        )
        for key in default_sequence
    ]
    module_type = str(module.get("type") or "fixed")
    if module_type not in MODULE_TYPES:
        module_type = "fixed"
    return ProjectModule(
        template_module_key=str(module["key"]),
        name=str(module.get("name") or module["key"]),
        module_type=module_type,
        page_templates=templates,
        default_sequence=default_sequence,
        default_add_template=default_add_template,
        slides=slides,
    )


def sync_legacy_module_state(project: PptProject) -> None:
    """Keep schema-v1 module fields useful without representing duplicate instances."""
    seen: set[str] = set()
    order: list[str] = []
    enabled: list[str] = []
    for module in project.modules:
        key = module.template_module_key
        if module.generated_by_binding_id or not key or key in seen:
            continue
        seen.add(key)
        order.append(key)
        if module.enabled:
            enabled.append(key)
    project.module_order = order
    project.enabled_modules = enabled


def ensure_project_modules(
    project: PptProject,
    manifest: "TemplateManifest",
    *,
    reset: bool = False,
) -> None:
    """Initialize schema-v2 modules or migrate a loaded schema-v1 project."""
    if project.modules and not reset:
        return
    legacy_order = list(project.module_order)
    legacy_enabled = set(project.enabled_modules)
    modules_by_key = {str(item["key"]): item for item in manifest.modules}
    ordered_keys = [key for key in legacy_order if key in modules_by_key]
    ordered_keys.extend(key for key in modules_by_key if key not in ordered_keys)
    project.modules = [
        _new_module_from_manifest(modules_by_key[key], manifest.slide_count)
        for key in ordered_keys
    ]
    if legacy_enabled:
        for module in project.modules:
            module.enabled = module.template_module_key in legacy_enabled
    project.module_bindings = [] if reset else project.module_bindings
    sync_legacy_module_state(project)


def module_by_id(project: PptProject, module_id: str) -> ProjectModule:
    for module in project.modules:
        if module.id == module_id:
            return module
    raise ValueError(f"项目中不存在模块：{module_id}")


def slide_by_id(module: ProjectModule, slide_id: str) -> ProjectSlide:
    for slide in module.slides:
        if slide.id == slide_id:
            return slide
    raise ValueError(f"模块 {module.name} 中不存在页面：{slide_id}")


def page_template_by_key(module: ProjectModule, key: str) -> PageTemplateRef:
    for page_template in module.page_templates:
        if page_template.key == key:
            return page_template
    raise ValueError(f"模块 {module.name} 中不存在页面模板：{key}")


def _unique_module_name(project: PptProject, base_name: str) -> str:
    existing = {item.name for item in project.modules}
    if base_name not in existing:
        return base_name
    index = 2
    while f"{base_name} {index}" in existing:
        index += 1
    return f"{base_name} {index}"


def add_module(
    project: PptProject,
    manifest: "TemplateManifest",
    template_module_key: str,
    *,
    name: str = "",
    position: int | None = None,
) -> ProjectModule:
    raw = next(
        (item for item in manifest.modules if item["key"] == template_module_key),
        None,
    )
    if raw is None:
        raise ValueError(f"模板中不存在模块：{template_module_key}")
    module = _new_module_from_manifest(raw, manifest.slide_count)
    module.name = _unique_module_name(project, name.strip() or module.name)
    index = len(project.modules) if position is None else max(0, min(position, len(project.modules)))
    project.modules.insert(index, module)
    sync_legacy_module_state(project)
    return module


def duplicate_module(project: PptProject, module_id: str) -> ProjectModule:
    source = module_by_id(project, module_id)
    duplicate = deepcopy(source)
    duplicate.id = new_project_id()
    duplicate.name = _unique_module_name(project, f"{source.name} - 副本")
    duplicate.generated_by_binding_id = ""
    duplicate.generated_row_index = None
    duplicate.enabled = True
    for slide in duplicate.slides:
        slide.id = new_project_id()
    index = project.modules.index(source) + 1
    project.modules.insert(index, duplicate)
    sync_legacy_module_state(project)
    return duplicate


def remove_module(project: PptProject, module_id: str) -> ProjectModule:
    module = module_by_id(project, module_id)
    binding_ids = {
        item.id for item in project.module_bindings if item.source_module_id == module_id
    }
    removal_ids = {
        item.id
        for item in project.modules
        if item.id == module_id or item.generated_by_binding_id in binding_ids
    }
    if len(project.modules) - len(removal_ids) < 1:
        raise ValueError("项目至少需要保留一个模块")
    project.modules = [item for item in project.modules if item.id not in removal_ids]
    project.module_bindings = [
        item for item in project.module_bindings if item.source_module_id != module_id
    ]
    sync_legacy_module_state(project)
    return module


def move_module(project: PptProject, module_id: str, offset: int) -> bool:
    module = module_by_id(project, module_id)
    current = project.modules.index(module)
    target = current + offset
    if target < 0 or target >= len(project.modules):
        return False
    project.modules.pop(current)
    project.modules.insert(target, module)
    sync_legacy_module_state(project)
    return True


def add_page_template(
    module: ProjectModule,
    source_slide_number: int,
    name: str,
    *,
    role: str = "content",
) -> PageTemplateRef:
    if source_slide_number < 1:
        raise ValueError("页面模板来源页必须为正整数")
    page_template = PageTemplateRef(
        key=f"custom_page_{source_slide_number}_{new_project_id()[:8]}",
        name=name.strip() or f"自定义模板第{source_slide_number}页",
        source_slide=source_slide_number,
        role=role if role in PAGE_ROLES else "content",
        default_title=name.strip(),
    )
    module.page_templates.append(page_template)
    if not module.default_sequence:
        module.default_sequence = [page_template.key]
    if not module.default_add_template:
        module.default_add_template = page_template.key
    return page_template


def set_default_page_template(module: ProjectModule, page_template_key: str) -> None:
    page_template_by_key(module, page_template_key)
    module.default_add_template = page_template_key


def add_slide(
    module: ProjectModule,
    page_template_key: str = "",
    *,
    position: int | None = None,
) -> ProjectSlide:
    key = page_template_key or module.default_add_template
    if not key:
        raise ValueError("当前模块没有默认新增页面模板，请先选择页面模板")
    page_template = page_template_by_key(module, key)
    slide = ProjectSlide(
        page_template_key=key,
        title=page_template.default_title or page_template.name,
    )
    index = len(module.slides) if position is None else max(0, min(position, len(module.slides)))
    module.slides.insert(index, slide)
    return slide


def duplicate_slide(module: ProjectModule, slide_id: str) -> ProjectSlide:
    source = slide_by_id(module, slide_id)
    duplicate = deepcopy(source)
    duplicate.id = new_project_id()
    base_title = source.title or page_template_by_key(module, source.page_template_key).name
    existing = {item.title for item in module.slides}
    index = 2
    suggested = f"{base_title} {index}"
    while suggested in existing:
        index += 1
        suggested = f"{base_title} {index}"
    duplicate.title = suggested
    position = module.slides.index(source) + 1
    module.slides.insert(position, duplicate)
    return duplicate


def remove_slide(module: ProjectModule, slide_id: str) -> ProjectSlide:
    if len(module.slides) <= 1:
        raise ValueError("模块至少需要保留一页；如不需要该模块请删除整个模块")
    slide = slide_by_id(module, slide_id)
    module.slides.remove(slide)
    return slide


def move_slide(module: ProjectModule, slide_id: str, offset: int) -> bool:
    slide = slide_by_id(module, slide_id)
    current = module.slides.index(slide)
    target = current + offset
    if target < 0 or target >= len(module.slides):
        return False
    module.slides.pop(current)
    module.slides.insert(target, slide)
    return True


def read_excel_records(binding: ExcelModuleBinding) -> list[dict[str, str]]:
    source = Path(binding.source_path).expanduser().resolve()
    if source.suffix.lower() not in {".xlsx", ".xlsm"} or not source.is_file():
        raise FileNotFoundError(f"模块绑定 Excel 不存在或格式不支持：{source}")
    if binding.header_row < 1:
        raise ValueError("模块绑定表头行必须大于 0")
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if binding.sheet not in workbook.sheetnames:
            raise ValueError(f"Excel 中不存在工作表：{binding.sheet}")
        worksheet = workbook[binding.sheet]
        min_column = 1
        max_column = int(worksheet.max_column or 0)
        data_start_row = binding.header_row + 1
        max_row = int(worksheet.max_row or 0)
        if binding.data_range:
            min_column, range_min_row, max_column, max_row = range_boundaries(
                binding.data_range
            )
            data_start_row = max(data_start_row, range_min_row)
        if not max_column or not max_row:
            return []

        header_values = next(
            worksheet.iter_rows(
                min_row=binding.header_row,
                max_row=binding.header_row,
                min_col=min_column,
                max_col=max_column,
                values_only=True,
            )
        )
        headers = [excel_value_text(value).strip() for value in header_values]
        nonempty_headers = [item for item in headers if item]
        if not nonempty_headers:
            raise ValueError("Excel 表头行没有可用字段")
        if len(nonempty_headers) != len(set(nonempty_headers)):
            raise ValueError("Excel 表头包含重复字段，请先修改表头")

        records: list[dict[str, str]] = []
        for row in worksheet.iter_rows(
            min_row=data_start_row,
            max_row=max_row,
            min_col=min_column,
            max_col=max_column,
            values_only=True,
        ):
            values = [excel_value_text(value) for value in row]
            record = {
                header: values[index]
                for index, header in enumerate(headers)
                if header
            }
            if any(value.strip() for value in record.values()):
                records.append(record)
        return records
    finally:
        workbook.close()


def materialize_excel_modules(
    project: PptProject,
    binding: ExcelModuleBinding,
) -> list[ProjectModule]:
    source_module = module_by_id(project, binding.source_module_id)
    if source_module.generated_by_binding_id:
        raise ValueError("Excel 绑定源不能是数据生成的模块副本")
    records = read_excel_records(binding)
    project.modules = [
        module
        for module in project.modules
        if module.generated_by_binding_id != binding.id
    ]
    source_module = module_by_id(project, binding.source_module_id)
    source_module.module_type = "dataDriven"
    source_module.enabled = False

    generated: list[ProjectModule] = []
    for row_index, record in enumerate(records, start=1):
        duplicate = deepcopy(source_module)
        duplicate.id = new_project_id()
        duplicate.enabled = True
        duplicate.generated_by_binding_id = binding.id
        duplicate.generated_row_index = row_index
        mapped_values = {
            target_key: record.get(source_header, "")
            for source_header, target_key in binding.field_map.items()
            if target_key
        }
        duplicate.module_values.update(mapped_values)
        name_value = record.get(binding.module_name_field, "").strip()
        duplicate.name = (
            f"{source_module.name} · {name_value}"
            if name_value
            else f"{source_module.name} · {row_index}"
        )
        for slide in duplicate.slides:
            slide.id = new_project_id()
        generated.append(duplicate)

    insert_at = project.modules.index(source_module) + 1
    for offset, module in enumerate(generated):
        project.modules.insert(insert_at + offset, module)
    existing_index = next(
        (index for index, item in enumerate(project.module_bindings) if item.id == binding.id),
        None,
    )
    if existing_index is None:
        project.module_bindings.append(binding)
    else:
        project.module_bindings[existing_index] = binding
    sync_legacy_module_state(project)
    return generated


def rebuild_structure_context(
    project: PptProject,
    manifest: "TemplateManifest",
) -> list[SlideContext]:
    enabled_modules = [module for module in project.modules if module.enabled and module.slides]
    physical_total = sum(len(module.slides) for module in enabled_modules)
    total_pages = sum(
        1
        for module in enabled_modules
        for slide in module.slides
        if page_template_by_key(module, slide.page_template_key).role
        not in NUMBER_EXCLUDED_ROLES
    )
    module_count = len(enabled_modules)
    contexts: list[SlideContext] = []
    physical_number = 0
    content_number = 0
    for module_number, module in enumerate(enabled_modules, start=1):
        module_page_count = len(module.slides)
        for module_page_number, slide in enumerate(module.slides, start=1):
            physical_number += 1
            page_template = page_template_by_key(module, slide.page_template_key)
            page_number: int | None = None
            if page_template.role not in NUMBER_EXCLUDED_ROLES:
                content_number += 1
                page_number = content_number
            slide_title = (
                slide.title
                or str(module.module_values.get("slide_title") or "")
                or page_template.default_title
                or page_template.name
            )
            slide_subtitle = (
                slide.subtitle
                or str(module.module_values.get("slide_subtitle") or "")
                or page_template.default_subtitle
            )
            values = deepcopy(project.values)
            values.update(deepcopy(module.module_values))
            values.update(deepcopy(slide.overrides))
            system_values: dict[str, Any] = {
                "page_number": page_number if page_number is not None else "",
                "total_pages": total_pages,
                "physical_page_number": physical_number,
                "physical_total_pages": physical_total,
                "module_number": module_number,
                "module_count": module_count,
                "module_page_number": module_page_number,
                "module_page_count": module_page_count,
                "module_title": module.name,
                "slide_title": slide_title,
                "slide_subtitle": slide_subtitle,
            }
            try:
                page_number_text = page_template.number_format.format(**system_values)
            except (KeyError, ValueError) as exc:
                raise ValueError(
                    f"页面模板 {page_template.name} 的页码格式无效：{page_template.number_format}"
                ) from exc
            system_values["page_number_text"] = (
                page_number_text if page_number is not None else ""
            )
            values.update(system_values)
            contexts.append(
                SlideContext(
                    module_id=module.id,
                    slide_id=slide.id,
                    template_module_key=module.template_module_key,
                    page_template_key=page_template.key,
                    source_slide=page_template.source_slide,
                    role=page_template.role,
                    physical_page_number=physical_number,
                    physical_total_pages=physical_total,
                    page_number=page_number,
                    total_pages=total_pages,
                    module_number=module_number,
                    module_count=module_count,
                    module_page_number=module_page_number,
                    module_page_count=module_page_count,
                    module_title=module.name,
                    slide_title=slide_title,
                    slide_subtitle=slide_subtitle,
                    values=values,
                    system_slots=dict(page_template.system_slots),
                    remove_shape_ids=list(page_template.remove_shape_ids),
                    number_format=page_template.number_format,
                )
            )
    return contexts


def build_slide_instances(
    project: PptProject,
    manifest: "TemplateManifest",
) -> list[SlideContext]:
    """Compatibility name used by the approved Spec."""
    return rebuild_structure_context(project, manifest)


def slot_specs_for_source_slide(
    manifest: "TemplateManifest",
    source_slide: int,
) -> list[dict[str, Any]]:
    return [dict(slot) for slot in manifest.slots if slot["slide"] == source_slide]
