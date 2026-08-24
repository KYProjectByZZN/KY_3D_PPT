"""Build a generic Excel fixture for manually reviewing the mapping workbench."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def build(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "方案数据"
    sheet.append(["项目标题", "【Excel映射验证】筒形壳体视觉检测方案"])
    sheet.append(["项目编号", "编号：EXCEL-MAP-001"])
    sheet.append(["方案日期", "2026年08月"])
    sheet.append([])
    sheet.append(["设备参数表（选择 A6:D12）"])

    parameter_rows = [
        ["检测节拍", "36 pcs/min", "工业光源", "LED频闪光源"],
        ["设备电压", "220V/50Hz", "工业相机", "500万像素面阵"],
        ["设备功率", "3.0kW", "工业镜头", "低畸变FA镜头"],
        ["运行环境", "Windows 10/11", "数据接口", "MES预留"],
        ["视觉软件", "OpenexVision", "产品范围", "按封样确认"],
        ["图片格式", "PNG/JPG", "检测输出", "OK/NG及缺陷坐标"],
        ["数据格式", "CSV/XLSX", "工作环境", "室内常温"],
    ]
    for row in parameter_rows:
        sheet.append(row)

    sheet.append([])
    sheet.append(["检测项目表（选择 A15:E23）"])
    inspection_rows = [
        ["序号", "检测位置", "是否检测", "检测精度", "备注"],
        ["1", "A面压痕", "是", "0.8×0.8×0.5 mm", "外观"],
        ["2", "B面开裂", "是", "0.8×0.5×0.2 mm", "外观"],
        ["3", "B面台阶", "是", "0.8×0.5×0.2 mm", "尺寸"],
        ["4", "侧面压痕", "是", "0.8×0.8×0.5 mm", "旋转检测"],
        ["5", "C面毛刺", "是", "0.8×0.5×0.2 mm", "旋转检测"],
        ["6", "C面磨痕", "是", "2×2 mm", "旋转检测"],
        ["7", "下孔异常", "是", "按样品确认", "底部检测"],
        ["8", "缺孔", "是", "有/无", "底部检测"],
    ]
    for row in inspection_rows:
        sheet.append(row)

    for cell in (sheet["A5"], sheet["A14"]):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for row_number in (6, 15):
        for cell in sheet[row_number]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DBEAFE")
    for column, width in {"A": 18, "B": 28, "C": 18, "D": 25, "E": 18}.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output",
        type=Path,
        default=PROJECT_ROOT / "examples" / "excel_mapping_demo.xlsx",
    )
    args = parser.parse_args()
    output = build(args.output.resolve())
    print(f"Excel demo: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
