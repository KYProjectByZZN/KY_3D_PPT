"""Build the reproducible Excel fixture for module-repeat demonstrations."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT = PROJECT_ROOT / "examples" / "module_repeat_demo.xlsx"


def main() -> int:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "设备模块"
    sheet.append(["设备标题", "设备说明"])
    sheet.append(["上料工位设备方案", "上料、定位与状态确认模块。"])
    sheet.append(["视觉检测工位方案", "相机、镜头和光源组成的视觉检测模块。"])
    sheet.append(["复检与下料工位方案", "复检、结果输出与自动下料模块。"])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 48
    sheet.freeze_panes = "A2"
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(OUTPUT.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
