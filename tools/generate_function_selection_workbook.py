from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_PATH = (
    Path(__file__).resolve().parents[1]
    / "output"
    / "本地AI_PPT生成软件功能选择清单.xlsx"
)

MVP_IDS = {"A1", "A4", "B1", "B2", "B3", "B6", "C1", "D1", "D2", "D3"}

FEATURES = [
    ("A1", "AI 内容功能", "输入项目需求，AI自动生成PPT大纲", "PPTAgent、Presenton", "低"),
    ("A2", "AI 内容功能", "AI从预设页面中选择需要哪些页面，如需求、方案、参数、案例", "PPTAgent", "中"),
    ("A3", "AI 内容功能", "AI为每一页选择合适Layout，如图文页、参数表、流程页", "PPTAgent", "中"),
    ("A4", "AI 内容功能", "AI生成每页标题、正文和要点", "PPTAgent、Presenton", "低"),
    ("A5", "AI 内容功能", "根据模板容量限制文字长度和要点数量", "PPTAgent", "中"),
    ("A6", "AI 内容功能", "导入Word、PDF等资料，AI提取内容生成PPT", "DeepPresenter、paper-ppt-agent", "中"),
    ("A7", "AI 内容功能", "AI判断某页是否需要图片、表格或流程图", "PPTAgent", "中"),
    ("A8", "AI 内容功能", "AI根据页面内容，从素材库选择匹配图片", "PPTAgent", "中"),
    ("B1", "模板与PPT生成", "用户选择一个现有PPT模板", "python-pptx、pptx-automizer", "低"),
    ("B2", "模板与PPT生成", "复制模板中的指定页面，保留原有设计", "pptx-automizer", "中"),
    ("B3", "模板与PPT生成", "根据形状名称替换标题、正文和图片", "pptx-automizer、pptx-template", "低"),
    ("B4", "模板与PPT生成", "自动识别模板中的Master、Layout和Placeholder", "PPTAgent、python-pptx", "高"),
    ("B5", "模板与PPT生成", "把一个PPT自动整理成可复用的Layout库", "PPTAgent、Presenton", "高"),
    ("B6", "模板与PPT生成", "输出文字、图片、表格可编辑的PPTX", "PptxGenJS、python-pptx", "中"),
    ("B7", "模板与PPT生成", "生成原生PowerPoint表格", "PptxGenJS、python-pptx", "中"),
    ("B8", "模板与PPT生成", "生成数据可编辑的PowerPoint图表", "PptxGenJS、pptx-automizer", "中"),
    ("B9", "模板与PPT生成", "内容太多时自动缩短、换Layout或拆成两页", "PPTAgent", "高"),
    ("B10", "模板与PPT生成", "检查PPTX是否损坏、图片是否缺失", "pptx-automizer、Open XML SDK", "中"),
    ("C1", "图片和图示", "自动插入用户提供的产品图、设备图", "PptxGenJS、python-pptx", "低"),
    ("C2", "图片和图示", "根据文字自动生成流程图", "Mermaid", "中"),
    ("C3", "图片和图示", "自动生成柱状图、折线图、饼图等", "PptxGenJS、ECharts", "中"),
    ("C4", "图片和图示", "调用AI生成示意图或背景图", "Presenton", "中"),
    ("C5", "图片和图示", "建立本地产品图片和案例图片素材库", "Presenton", "中"),
    ("D1", "本地软件交互", "输入需求、选择模板、点击生成", "Presenton 的流程设计", "低"),
    ("D2", "本地软件交互", "生成前查看并修改PPT大纲", "Presenton", "低"),
    ("D3", "本地软件交互", "生成前勾选需要的页面", "模块化方案设计", "低"),
    ("D4", "本地软件交互", "预览每一页生成效果", "Presenton、PPTist", "中"),
    ("D5", "本地软件交互", "单独重新生成某一页", "paper-ppt-agent", "中"),
    ("D6", "本地软件交互", "修改某页文字后重新生成PPTX", "PPTist", "中"),
    ("D7", "本地软件交互", "保存本次项目，下次继续修改", "paper-ppt-agent", "低"),
    ("D8", "本地软件交互", "选择不同AI模型或配置API Key", "Presenton", "低"),
    ("E1", "质量检查", "自动检查文字溢出", "paper-ppt-agent", "中"),
    ("E2", "质量检查", "自动检查元素重叠", "paper-ppt-agent", "中"),
    ("E3", "质量检查", "检查图片分辨率和拉伸", "Presenton、PPTist", "中"),
    ("E4", "质量检查", "AI查看生成后的页面图片并提出修改", "paper-ppt-agent", "高"),
    ("E5", "质量检查", "生成失败时自动换Layout重试", "PPTAgent", "高"),
]

REFERENCES = [
    ("PPTAgent / DeepPresenter", "https://github.com/icip-cas/PPTAgent", "大纲规划、模块选择、Layout选择、模板归纳", "适合借鉴AI编排方法，不建议把整个研究型运行时直接产品化"),
    ("Presenton", "https://github.com/presenton/presenton", "本地软件流程、模板工作台、编辑与预览交互", "借鉴产品流程；核心导出运行时需单独评估"),
    ("paper-ppt-agent", "https://github.com/CRui5in/paper-ppt-agent", "单页重生成、质量审查、项目保存", "AGPL，闭源软件中先作为架构参考"),
    ("PptxGenJS", "https://github.com/gitbrent/PptxGenJS", "原生可编辑文字、图片、表格、图表、形状", "适合作为新建PPTX对象的自动化引擎"),
    ("python-pptx", "https://github.com/scanny/python-pptx", "模板分析、Placeholder、轻量修改与结构检查", "可保留现有Python能力；复杂模板复制需配合其他方案"),
    ("pptx-automizer", "https://github.com/singerla/pptx-automizer", "复制现有模板页、保留设计、替换对象", "适合企业模板复用，但必须用真实模板做兼容性测试"),
    ("pptx-template", "https://github.com/m3dev/pptx-template", "按标记替换模板文字、表格和图表数据", "适合借鉴模板字段绑定方式"),
    ("PPTist", "https://github.com/pipipi-pikachu/PPTist", "页面预览、拖拽编辑、PPTX映射", "AGPL；优先参考交互与数据模型"),
    ("Open XML SDK", "https://github.com/dotnet/Open-XML-SDK", "PPTX包结构与关系验证", "适合作为可选验证工具，不作为主生成引擎"),
    ("Mermaid", "https://github.com/mermaid-js/mermaid", "文字生成流程图并导出SVG", "插入PPT后通常不是原生可编辑图形"),
    ("ECharts", "https://github.com/apache/echarts", "复杂图表与SVG输出", "原生PowerPoint图表优先，ECharts作为复杂图表补充"),
]

NAVY = "1F4E78"
BLUE = "D9EAF7"
PALE_GREEN = "E2F0D9"
PALE_YELLOW = "FFF2CC"
PALE_RED = "FCE4D6"
PALE_PURPLE = "E4DFEC"
WHITE = "FFFFFF"
GRID = Side(style="thin", color="D9E2F3")
FONT_NAME = "Microsoft YaHei"


def style_title(ws, title: str, subtitle: str, end_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT_NAME, size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT_NAME, size=10, color="44546A")
    ws["A2"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 34


def style_header_row(ws, row: int, end_column: int) -> None:
    for cell in ws.iter_cols(min_row=row, max_row=row, min_col=1, max_col=end_column):
        item = cell[0]
        item.font = Font(name=FONT_NAME, bold=True, color=WHITE)
        item.fill = PatternFill("solid", fgColor=NAVY)
        item.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        item.border = Border(left=GRID, right=GRID, top=GRID, bottom=GRID)
    ws.row_dimensions[row].height = 26


def build_selection_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "功能选择"
    ws.sheet_view.showGridLines = False
    style_title(
        ws,
        "本地 AI PPT 生成软件｜功能选择清单",
        "已从上一轮会话恢复原始36项功能。请在“我的选择”列使用下拉菜单；★/“是”表示上一轮建议的第一版核心功能。",
        8,
    )

    headers = ["编号", "功能分类", "功能说明", "参考项目", "实现难度", "第一版建议", "我的选择", "备注"]
    for column, value in enumerate(headers, start=1):
        ws.cell(row=4, column=column, value=value)
    style_header_row(ws, 4, len(headers))

    category_colors = {
        "AI 内容功能": "DDEBF7",
        "模板与PPT生成": "E2F0D9",
        "图片和图示": "FFF2CC",
        "本地软件交互": "E4DFEC",
        "质量检查": "FCE4D6",
    }

    for row_number, (feature_id, category, description, source, difficulty) in enumerate(FEATURES, start=5):
        values = [
            feature_id,
            category,
            description,
            source,
            difficulty,
            "是（★）" if feature_id in MVP_IDS else "否",
            "未选择",
            "",
        ]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row_number, column=column, value=value)
            cell.font = Font(name=FONT_NAME, size=10, bold=(column == 1))
            cell.alignment = Alignment(
                horizontal="center" if column in {1, 2, 5, 6, 7} else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.border = Border(left=GRID, right=GRID, top=GRID, bottom=GRID)
            if column == 2:
                cell.fill = PatternFill("solid", fgColor=category_colors[category])
            if column == 6 and feature_id in MVP_IDS:
                cell.fill = PatternFill("solid", fgColor=PALE_GREEN)
        ws.row_dimensions[row_number].height = 38

    final_row = 4 + len(FEATURES)
    table = Table(displayName="FunctionSelectionTable", ref=f"A4:H{final_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    validation = DataValidation(
        type="list",
        formula1='"未选择,保留（第一版）,保留（后续）,暂不需要,待讨论"',
        allow_blank=False,
    )
    validation.error = "请从下拉列表中选择。"
    validation.errorTitle = "选择无效"
    validation.prompt = "请选择：第一版、后续、暂不需要或待讨论。"
    validation.promptTitle = "功能选择"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    ws.add_data_validation(validation)
    validation.add(f"G5:G{final_row}")

    selection_range = f"G5:G{final_row}"
    ws.conditional_formatting.add(
        selection_range,
        CellIsRule(operator="equal", formula=['"保留（第一版）"'], fill=PatternFill("solid", fgColor=PALE_GREEN)),
    )
    ws.conditional_formatting.add(
        selection_range,
        CellIsRule(operator="equal", formula=['"保留（后续）"'], fill=PatternFill("solid", fgColor=BLUE)),
    )
    ws.conditional_formatting.add(
        selection_range,
        CellIsRule(operator="equal", formula=['"暂不需要"'], fill=PatternFill("solid", fgColor=PALE_RED)),
    )
    ws.conditional_formatting.add(
        selection_range,
        CellIsRule(operator="equal", formula=['"待讨论"'], fill=PatternFill("solid", fgColor=PALE_YELLOW)),
    )

    widths = {"A": 9, "B": 18, "C": 48, "D": 30, "E": 11, "F": 14, "G": 18, "H": 32}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:H{final_row}"
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("填写说明与汇总")
    ws.sheet_view.showGridLines = False
    style_title(
        ws,
        "填写说明与选择汇总",
        "产品边界：本地软件；通过AI或自动化代码生成PPT；本清单不要求建设前端/后端平台。",
        5,
    )

    ws["A4"] = "怎么填写"
    ws["A4"].font = Font(name=FONT_NAME, size=13, bold=True, color=NAVY)
    instructions = [
        "1. 打开“功能选择”工作表。",
        "2. 在G列“我的选择”中逐项选择：保留（第一版）/ 保留（后续）/ 暂不需要 / 待讨论。",
        "3. 有补充要求时写在H列“备注”。",
        "4. 保存文件后直接把这个Excel交回来，就可以据此整理正式需求和开发顺序。",
    ]
    for row, text in enumerate(instructions, start=5):
        ws.cell(row=row, column=1, value=text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=10)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 24

    ws["A10"] = "动态汇总（打开Excel后自动计算）"
    ws["A10"].font = Font(name=FONT_NAME, size=13, bold=True, color=NAVY)
    summary = [
        ("功能总数", "=COUNTA('功能选择'!A5:A40)"),
        ("上一轮第一版建议", '=COUNTIF(\'功能选择\'!F5:F40,"是（★）")'),
        ("已选：第一版", '=COUNTIF(\'功能选择\'!G5:G40,"保留（第一版）")'),
        ("已选：后续", '=COUNTIF(\'功能选择\'!G5:G40,"保留（后续）")'),
        ("暂不需要", '=COUNTIF(\'功能选择\'!G5:G40,"暂不需要")'),
        ("待讨论", '=COUNTIF(\'功能选择\'!G5:G40,"待讨论")'),
        ("未选择", '=COUNTIF(\'功能选择\'!G5:G40,"未选择")'),
    ]
    for row, (label, formula) in enumerate(summary, start=11):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=formula)
        for column in range(1, 3):
            cell = ws.cell(row=row, column=column)
            cell.font = Font(name=FONT_NAME, bold=(column == 1))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=GRID, right=GRID, top=GRID, bottom=GRID)
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=BLUE)

    ws["A20"] = "上一轮建议的最小版本（10项）"
    ws["A20"].font = Font(name=FONT_NAME, size=13, bold=True, color=NAVY)
    recommended = [item for item in FEATURES if item[0] in MVP_IDS]
    for row, item in enumerate(recommended, start=21):
        ws.cell(row=row, column=1, value=item[0])
        ws.cell(row=row, column=2, value=item[2])
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        for column in range(1, 6):
            cell = ws.cell(row=row, column=column)
            cell.font = Font(name=FONT_NAME, size=10, bold=(column == 1))
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = Border(left=GRID, right=GRID, top=GRID, bottom=GRID)
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=PALE_GREEN)
        ws.row_dimensions[row].height = 24

    widths = {"A": 24, "B": 24, "C": 24, "D": 24, "E": 24}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A4"


def build_reference_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("参考项目说明")
    ws.sheet_view.showGridLines = False
    style_title(
        ws,
        "参考项目与可借鉴能力",
        "这些项目是功能来源，不代表整套采用；商业软件仍需结合许可证、可维护性和本地稳定性选择。",
        4,
    )
    headers = ["参考项目", "GitHub地址", "可借鉴能力", "使用边界"]
    for column, value in enumerate(headers, start=1):
        ws.cell(row=4, column=column, value=value)
    style_header_row(ws, 4, len(headers))

    for row, values in enumerate(REFERENCES, start=5):
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=column, value=value)
            cell.font = Font(name=FONT_NAME, size=10, color="0563C1" if column == 2 else "000000")
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = Border(left=GRID, right=GRID, top=GRID, bottom=GRID)
            if column == 2:
                cell.hyperlink = value
                cell.style = "Hyperlink"
        ws.row_dimensions[row].height = 44

    final_row = 4 + len(REFERENCES)
    table = Table(displayName="ReferenceProjectTable", ref=f"A4:D{final_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    widths = {"A": 26, "B": 52, "C": 42, "D": 52}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def validate_workbook(path: Path) -> None:
    with path.open("rb") as handle:
        if handle.read(4) != b"PK\x03\x04":
            raise RuntimeError("XLSX ZIP signature is invalid")

    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != ["功能选择", "填写说明与汇总", "参考项目说明"]:
        raise RuntimeError(f"Unexpected sheets: {workbook.sheetnames}")
    selection = workbook["功能选择"]
    if selection.max_row != 40 or selection["A5"].value != "A1" or selection["A40"].value != "E5":
        raise RuntimeError("Feature rows are incomplete")
    if len(selection.data_validations.dataValidation) != 1:
        raise RuntimeError("Selection dropdown validation is missing")
    if "FunctionSelectionTable" not in selection.tables:
        raise RuntimeError("Feature table is missing")


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.properties.creator = "Codex"
    workbook.properties.title = "本地AI PPT生成软件功能选择清单"
    workbook.properties.subject = "从上一轮会话恢复的36项功能选择菜单"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    build_selection_sheet(workbook)
    build_summary_sheet(workbook)
    build_reference_sheet(workbook)
    workbook.active = 0
    workbook.save(OUTPUT_PATH)
    validate_workbook(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
